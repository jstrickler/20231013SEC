"""
Copy a range from an existing spreadsheet
to a new sheet
"""
import openpyxl as px

WB_NAME = '../DATA/presidents.xlsx'

def main():
    """program entry point"""
    wb = px.load_workbook(WB_NAME)
    ws = wb['US Presidents']
    save_range_to_new_worksheet(wb, ws, 'President Names', 'B2','C47')
    wb.save(WB_NAME)

def save_range_to_new_worksheet(wb, old_ws, sheet_title, range_start, range_end):
    """Print first and last names of all presidents"""

    new_ws = wb.create_sheet(sheet_title)

    for row in old_ws[range_start:range_end]:
        for col in row:
            # new_ws.cell(row=x, column=y).value = col.value
            new_ws[col.coordinate] = col.value

if __name__ == '__main__':
    main()
