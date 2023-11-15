"""
Copy a range from an existing spreadsheet
to a new workbook
"""
import openpyxl as px

WB = '../DATA/presidents.xlsx'

SELECTORS = [
    ('A', 'alice'),
    ('D', 'betty'),
    ('F', 'charlotte'),
]

def main():
    """program entry point"""
    wb = px.load_workbook(WB)
    ws = wb['US Presidents']

    save_ranges_to_new_workbooks(ws)

    wb.close()

# "save as":
#    wb.save("presidents4.xlsx")

def save_ranges_to_new_workbooks(old_ws):
    for column_letter, user_name in SELECTORS:
        # copy column to new workbook
        wb = px.Workbook()
        new_ws = wb.active
        new_ws.title = "column {}".format(column_letter)

        range_start = f"{column_letter}1"
        range_end = f"{column_letter}{old_ws.max_row}"

        for row in old_ws[range_start:range_end]:
            for col in row:
                new_ws[col.coordinate] = col.value

        wb_name = f"{user_name}.xlsx"
        wb.save(wb_name)


if __name__ == '__main__':
    main()
