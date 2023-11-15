import openpyxl as px

def main():
    """program entry point"""
    wb = px.load_workbook('../DATA/presidents.xlsx')
    ws = wb['US Presidents']

    add_age_at_inauguration(ws)

    wb.save('presidents_formula.xlsx')

def add_age_at_inauguration(ws):
    """Add a new column with age of inauguration"""
    new_col = ws.max_column + 1
    print(new_col)
    ws.cell(row=1, column=new_col).value = 'Age at Inauguration'
    for row in range(2, ws.max_row + 1):
        new_cell = ws.cell(row=row, column=new_col)
        new_cell.value = f'=(H{row}-D{row})/365.25'
        new_cell.number_format = '0.0'


if __name__ == '__main__':
    main()
