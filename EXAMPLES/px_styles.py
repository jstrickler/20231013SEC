import openpyxl as px
from openpyxl.styles.fonts import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.colors import Color

def main():
    """program entry point"""
    wb = px.load_workbook('../DATA/presidents.xlsx')
    ws = wb['US Presidents']

    update_last_names(ws)
    add_border_to_state_column(ws)
    add_background_to_parties(ws)

    wb.save('presidents_styles.xlsx')

def add_border_to_state_column(ws):
    """Add a border around cells for the State column"""
    for row in ws['G2:G47']:
        cell = row[0]
        side = Side(style='thin')
        border = Border(
            left=side,
            right=side,
            top=side,
            bottom=side,
        )
        cell.border = border

def add_background_to_parties(ws):
    for row in ws['J2:J47']:
        cell = row[0]
        blue_bg = Color(rgb='006666FF', tint=.6)
        fill = PatternFill(
            patternType='solid',
            fgColor=blue_bg,
        )
        cell.fill = fill

def update_last_names(ws):
    """Make the last name column blue and bold"""
    for row in ws['B2:B47']:
        cell = row[0]
        cell.value = cell.value.upper()
        cell.font = Font(color='FF0000FF', name="Comic Sans", size=22)

if __name__ == '__main__':
    main()
